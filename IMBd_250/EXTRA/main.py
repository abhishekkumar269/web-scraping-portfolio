from bs4 import BeautifulSoup

import requests

import openpyxl

excel = openpyxl.Workbook()

print(excel.sheetnames)

sheet = excel.active

sheet.title = "Top rated movies"

print(excel.sheetnames)

sheet.append(['movie rank','movie name','ratings','year'])

try:

    headers = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/113.0.0.0 Safari/537.36"
}

    source = requests.get('https://www.imdb.com/chart/top/',headers=headers)
    source.raise_for_status()


    soup = BeautifulSoup(source.text,'html.parser')
    # print(soup)

    movies = soup.find('ul',class_="ipc-metadata-list ipc-metadata-list--dividers-between sc-e22973a9-0 khSCXM compact-list-view ipc-metadata-list--base").find_all('li')
    # print(len(movies))

    for movie in movies:


        name = movie.find("h3",class_= "ipc-title__text").text.split(".")[1]

        year = movie.find("span",class_="sc-4b408797-8 iurwGb cli-title-metadata-item").text

        ratings = movie.find("span" ,class_="ipc-rating-star--rating").text

        rank = movie.find("h3",class_="ipc-title__text").text.split('.')[0]

        sheet.append([rank,name,year,ratings])

except Exception as e:
    print(e)

excel.save("top 250.xlsx")